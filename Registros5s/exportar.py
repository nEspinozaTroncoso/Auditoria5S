import os
from flask import Blueprint, current_app, send_file, url_for, Response
from .models import Auditoria, Respuesta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from openpyxl.drawing.image import Image as XLImage
import webbrowser

try:
    import webview

    if webview.windows:
        window = webview.windows[0]
    else:
        window = None
except Exception:
    window = None

bp = Blueprint("exportar", __name__, url_prefix="/exportar")


@bp.route("/servir_excel/<int:auditoria_id>")
def servir_excel(auditoria_id):
    auditoria = Auditoria.query.get_or_404(auditoria_id)
    respuestas = Respuesta.query.filter_by(auditoria_id=auditoria.id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auditoría 5S"

    # --- Logo ---
    logo_path = os.path.join(current_app.static_folder, "img", "Sun_logo.png")
    if os.path.exists(logo_path):
        img_logo = XLImage(logo_path)
        img_logo.width = 180
        img_logo.height = 126
        ws.add_image(img_logo, "B2")
        ws.row_dimensions[2].height = 90

    # --- Título y datos principales ---
    ws.merge_cells("B3:D3")
    cell_title = ws["B3"]
    cell_title.value = "Auditoría 5S"
    cell_title.font = Font(bold=True, size=18)
    cell_title.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B4:D4")
    cell_area = ws["B4"]
    cell_area.value = f"Área: {auditoria.area}"
    cell_area.font = Font(size=12)
    cell_area.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B5:D5")
    cell_fecha = ws["B5"]
    cell_fecha.value = f"Fecha: {auditoria.fecha.strftime('%d-%m-%Y')}"
    cell_fecha.font = Font(size=12)
    cell_fecha.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B6:D6")
    cell_resp = ws["B6"]
    cell_resp.value = f"Responsable: {auditoria.responsable}"
    cell_resp.font = Font(size=12)
    cell_resp.alignment = Alignment(horizontal="center", vertical="center")

    start_row = 8  # Comienza después del logo y los datos principales
    start_col = 2  # Columna B

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    section_fill = PatternFill("solid", fgColor="A9D08E")
    section_font = Font(bold=True, color="000000")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    secciones_respuestas = {}
    for respuesta in respuestas:
        secciones_respuestas.setdefault(respuesta.seccion, []).append(respuesta)

    row_num = start_row
    img_col = start_col + 2  # Columna D

    for seccion, respuestas_seccion in secciones_respuestas.items():
        # Fila de sección
        ws.merge_cells(
            start_row=row_num,
            start_column=start_col,
            end_row=row_num,
            end_column=start_col + 2,
        )
        cell = ws.cell(row=row_num, column=start_col)
        cell.value = seccion
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = center_align
        cell.border = thin_border
        for col in range(start_col + 1, start_col + 3):
            merged_cell = ws.cell(row=row_num, column=col)
            merged_cell.font = section_font
            merged_cell.fill = section_fill
            merged_cell.alignment = center_align
            merged_cell.border = thin_border
        row_num += 1

        # Encabezados
        headers = ["Pregunta", "Puntaje (%)", "Imagen"]
        for col_num, header in enumerate(headers, start=start_col):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row_num += 1

        # Datos de la sección
        seccion_puntos = 0
        seccion_max = len(respuestas_seccion) * 100
        for respuesta in respuestas_seccion:
            cell_pregunta = ws.cell(
                row=row_num, column=start_col, value=respuesta.pregunta
            )
            cell_pregunta.alignment = center_align
            cell_pregunta.border = thin_border

            percent = f"{respuesta.puntaje}%" if respuesta.puntaje is not None else ""
            cell_puntaje = ws.cell(row=row_num, column=start_col + 1, value=percent)
            cell_puntaje.alignment = center_align
            cell_puntaje.border = thin_border

            seccion_puntos += respuesta.puntaje if respuesta.puntaje else 0

            # Insertar imagen si existe
            if respuesta.imagen_path:
                # 1. OBTENER LA RUTA BASE EXTERNA DESDE LA CONFIGURACIÓN
                # La variable 'UPLOAD_FOLDER' de Flask apunta a la ruta persistente externa (ej: .../Auditoria_5S_Datos/uploads)
                UPLOAD_DIR_EXTERNO = current_app.config["UPLOAD_FOLDER"]

                # 2. CONSTRUIR LA RUTA COMPLETA DE LA IMAGEN
                # respuesta.imagen_path solo contiene el nombre del archivo (ej: 'foto.jpg')
                img_path = os.path.join(
                    UPLOAD_DIR_EXTERNO,  # Usamos la ruta externa aquí
                    respuesta.imagen_path,
                )

                # El resto de la lógica de exportación permanece igual
                if os.path.exists(img_path):
                    # Asegúrate de que XLImage sea la clase correcta para la librería que usas (ej: openpyxl.drawing.image.Image)
                    img = XLImage(img_path)
                    img.width = 100
                    img.height = 100
                    img_cell = f"{get_column_letter(img_col)}{row_num}"
                    ws.add_image(img, img_cell)
                    ws.row_dimensions[row_num].height = 80

                ws.cell(row=row_num, column=img_col).border = thin_border
            else:
                ws.cell(row=row_num, column=img_col, value="").border = thin_border

            row_num += 1

        # Porcentaje total de la sección
        seccion_porcentaje = (
            round((seccion_puntos / seccion_max) * 100, 2) if seccion_max else 0
        )
        ws.merge_cells(
            start_row=row_num,
            start_column=start_col,
            end_row=row_num,
            end_column=start_col + 2,
        )
        for col in range(start_col, start_col + 3):
            cell_seccion_total = ws.cell(row=row_num, column=col)
            if col == start_col:
                cell_seccion_total.value = (
                    f"Porcentaje total de la sección: {seccion_porcentaje}%"
                )
            cell_seccion_total.font = Font(bold=True)
            cell_seccion_total.alignment = center_align
            cell_seccion_total.border = thin_border
        row_num += 2  # Fila vacía entre secciones

    # Porcentaje total al final
    ws.merge_cells(
        start_row=row_num,
        start_column=start_col,
        end_row=row_num,
        end_column=start_col + 2,
    )
    for col in range(start_col, start_col + 3):
        cell_total = ws.cell(row=row_num, column=col)
        if col == start_col:
            cell_total.value = f"Porcentaje Total General: {auditoria.total}%"
        cell_total.font = Font(bold=True, size=16)
        cell_total.alignment = center_align
        cell_total.border = thin_border

    # Ajustar ancho de columnas
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 18

    # Guardar en memoria y enviar
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    nombre_archivo = f"auditoria_{auditoria.id}.xlsx"
    return send_file(
        output,
        download_name=nombre_archivo,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# --- Función de Manejo (Abre el navegador y lo redirecciona) ---
@bp.route("/exportar_excel/<int:auditoria_id>")
def exportar_excel(auditoria_id):
    # Obtener el puerto y host actual (no los usaremos para concatenar, solo para contexto si es necesario)
    host = current_app.config.get("HOST", "127.0.0.1")
    port = current_app.config.get("PORT", 5000)

    # 1. CONSTRUCCIÓN CLAVE: Usar _external=True para obtener la URL COMPLETA
    # Flask debería generar: http://127.0.0.1:5000/exportar/servir_excel/1
    full_url = url_for(
        "exportar.servir_excel", auditoria_id=auditoria_id, _external=True
    )
    # 2. Abrir el navegador por defecto del sistema
    webbrowser.open_new_tab(full_url)
    # 3. Finalizar el hilo de PyWebView.
    return Response(status=200)
