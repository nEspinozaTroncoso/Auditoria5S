from flask import Blueprint, request, render_template, url_for, current_app, send_file
from .models import Auditoria, Respuesta
from .forms import secciones
from Registros5s import db
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from openpyxl.drawing.image import Image as XLImage

bp = Blueprint("registro", __name__, url_prefix="/registro")


@bp.route("/formulario", methods=("GET", "POST"))
def formulario():
    if request.method == "POST":
        responsable = request.form.get("responsable", "No indicado")
        resultados = {}
        total_puntos = 0
        total_max = 0

        # Crear auditoria
        auditoria = Auditoria(responsable=responsable, total=0)
        db.session.add(auditoria)
        db.session.commit()

        respuestas = []

        for seccion, preguntas in secciones.items():
            seccion_puntos = 0
            seccion_max = len(preguntas) * 100  # Cada pregunta puede valer hasta 100

            for idx, pregunta_obj in enumerate(preguntas):
                pregunta = pregunta_obj["pregunta"]
                opciones = pregunta_obj["opciones"]

                key = f"{seccion}_{idx}"
                valor = int(
                    request.form.get(key, 0)
                )  # El value será el valor ponderado
                seccion_puntos += valor

                # Guardar imagen
                imagen = request.files.get(f"{key}_img")
                imagen_path = None
                if imagen and imagen.filename != "":
                    if not os.path.exists(current_app.config["UPLOAD_FOLDER"]):
                        os.makedirs(current_app.config["UPLOAD_FOLDER"])
                    imagen_path = os.path.join(
                        current_app.config["UPLOAD_FOLDER"], imagen.filename
                    )
                    imagen_path = imagen_path.replace("\\", "/")

                    imagen.save(imagen_path)

                # Guardar respuesta en DB
                respuesta = Respuesta(
                    auditoria_id=auditoria.id,
                    seccion=seccion,
                    pregunta=pregunta,
                    puntaje=valor,
                    imagen_path=imagen_path,
                )
                db.session.add(respuesta)
                respuestas.append(respuesta)
            resultados[seccion] = round((seccion_puntos / seccion_max) * 100, 2)
            total_puntos += seccion_puntos
            total_max += seccion_max

        # Guardar total en auditoria
        auditoria.total = round((total_puntos / total_max) * 100, 2)
        db.session.commit()

        return render_template(
            "registro/resultado.html",
            resultados=resultados,
            total=auditoria.total,
            respuestas=respuestas,
        )
    return render_template("registro/formulario.html", secciones=secciones)


@bp.route("/historial")
def historial():
    page = request.args.get("page", 1, type=int)
    fecha_filtro = request.args.get("fecha", "")

    query = Auditoria.query
    if fecha_filtro:
        try:
            fecha_dt = datetime.strptime(fecha_filtro, "%Y-%m-%d")
            query = query.filter(db.func.date(Auditoria.fecha) == fecha_dt.date())
        except ValueError:
            pass

    auditorias = query.order_by(Auditoria.fecha.desc()).paginate(page=page, per_page=20)
    return render_template(
        "registro/historial.html",
        auditorias=auditorias,
        fecha_filtro=fecha_filtro,
    )


@bp.route("/detalle/<int:auditoria_id>")
def detalle(auditoria_id):
    auditoria = Auditoria.query.get_or_404(auditoria_id)
    respuestas = Respuesta.query.filter_by(auditoria_id=auditoria.id).all()

    # Agrupar respuestas por sección
    secciones_respuestas = {}
    for respuesta in respuestas:
        secciones_respuestas.setdefault(respuesta.seccion, []).append(respuesta)

    return render_template(
        "registro/detalle.html",
        auditoria=auditoria,
        secciones_respuestas=secciones_respuestas,
    )


@bp.route("/exportar_excel/<int:auditoria_id>")
def exportar_excel(auditoria_id):
    auditoria = Auditoria.query.get_or_404(auditoria_id)
    respuestas = Respuesta.query.filter_by(auditoria_id=auditoria.id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auditoría 5S"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    section_fill = PatternFill("solid", fgColor="A9D08E")
    section_font = Font(bold=True, color="000000")
    center_align = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    secciones_respuestas = {}
    for respuesta in respuestas:
        secciones_respuestas.setdefault(respuesta.seccion, []).append(respuesta)

    row_num = 1
    img_col = 3  # Columna donde van las imágenes

    for seccion, respuestas_seccion in secciones_respuestas.items():
        # Fila de sección
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
        cell = ws.cell(row=row_num, column=1)
        cell.value = seccion
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = center_align
        cell.border = thin_border
        row_num += 1

        # Encabezados
        headers = ["Pregunta", "Puntaje (%)", "Imagen"]
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row_num += 1

        # Datos de la sección
        for respuesta in respuestas_seccion:
            ws.cell(row=row_num, column=1, value=respuesta.pregunta).border = (
                thin_border
            )
            percent = f"{respuesta.puntaje}%" if respuesta.puntaje is not None else ""
            cell_puntaje = ws.cell(row=row_num, column=2, value=percent)
            cell_puntaje.alignment = center_align
            cell_puntaje.border = thin_border

            # Insertar imagen si existe
            if respuesta.imagen_path:
                img_path = os.path.join(
                    current_app.static_folder, respuesta.imagen_path
                )
                if os.path.exists(img_path):
                    img = XLImage(img_path)
                    img.width = 100  # Ajusta el tamaño de la imagen
                    img.height = 100
                    img_cell = f"{get_column_letter(img_col)}{row_num}"
                    ws.add_image(img, img_cell)
                    ws.row_dimensions[row_num].height = (
                        80  # Ajusta la altura de la fila
                    )
                ws.cell(row=row_num, column=img_col).border = thin_border
            else:
                ws.cell(row=row_num, column=img_col, value="").border = thin_border

            row_num += 1

        # Fila vacía entre secciones
        row_num += 1

    # Porcentaje total al final
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
    cell_total = ws.cell(row=row_num, column=1)
    cell_total.value = f"Porcentaje Total: {auditoria.total}%"
    cell_total.font = Font(bold=True, size=16)
    cell_total.alignment = center_align
    cell_total.border = thin_border

    # Ajustar ancho de columnas
    ws.column_dimensions["A"].width = 50  # Pregunta
    ws.column_dimensions["B"].width = 15  # Puntaje
    ws.column_dimensions["C"].width = 18  # Imagen

    # Guardar en memoria y enviar
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    nombre_archivo = f"auditoria_{auditoria.id}.xlsx"
    return send_file(
        output,
        download_name=nombre_archivo,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
